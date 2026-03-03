from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import os
import re
from io import BytesIO
import json

app = Flask(__name__)

def smart_read_excel(file_stream):
    df_raw = pd.read_excel(file_stream, header=None, nrows=15)
    header_row = 0
    for i, row in df_raw.iterrows():
        if row.notna().sum() >= 3:
            header_row = i
            break
            
    file_stream.seek(0)
    df = pd.read_excel(file_stream, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    return df.loc[:, ~df.columns.str.contains('^Unnamed')]

# --- NEW DATE PARSING LOGIC ---
def detect_hour(filename):
    date_match = re.search(r'(\d{8})_(\d{4})', filename)
    if date_match:
        date_str = date_match.group(1) 
        time_str = date_match.group(2) 
        try:
            dt_obj = datetime.strptime(f"{date_str}_{time_str}", "%Y%m%d_%H%M")
            return dt_obj, dt_obj.strftime("%b %d - %I:%M %p")
        except ValueError:
            pass

    time_match = re.search(r'_(\d{4})|(\d{4})', filename)
    if time_match:
        t = time_match.group(1) or time_match.group(2)
        try:
            dt_obj = datetime.strptime(t, "%H%M")
            today = datetime.today()
            dt_obj = dt_obj.replace(year=today.year, month=today.month, day=today.day)
            return dt_obj, dt_obj.strftime("%I:%M %p")
        except ValueError:
            pass
            
    return None, None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/get_columns', methods=['POST'])
def get_columns():
    if 'files' not in request.files:
        return jsonify({"error": "No files uploaded"}), 400
    try:
        df = smart_read_excel(request.files.getlist('files')[0])
        return jsonify({"columns": list(df.columns)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/process', methods=['POST'])
def process_files():
    if 'files' not in request.files:
        return jsonify({"error": "No files uploaded"}), 400

    files = request.files.getlist('files')
    primary_col = request.form.get('p_col')
    dt_col = request.form.get('dt_col')
    extra_cols = json.loads(request.form.get('extra_cols', '[]'))
    
    if not primary_col or not dt_col:
        return jsonify({"error": "Missing column selections"}), 400

    today = datetime.today().date()
    hourly_data = []

    try:
        for file in files:
            dt_obj, display_hour = detect_hour(file.filename)
            if dt_obj is None:
                continue

            df = smart_read_excel(file)
            df[dt_col] = pd.to_datetime(df[dt_col], errors="coerce")
            
            syncing = df[df[dt_col].dt.date == today].copy()
            unsync = df[df[dt_col].dt.date != today].copy()

            hourly_data.append({
                "hour": dt_obj, # Sort by actual datetime
                "display": display_hour, 
                "sync": syncing, 
                "unsync": unsync, 
                "raw": df
            })

        hourly_data.sort(key=lambda x: x["hour"])

        wb = Workbook()
        ws = wb.active
        ws.freeze_panes = "A2" # Freeze header
        
        sync_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        unsync_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        new_gp_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid") # New GP Yellow
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        border = Border(left=Side(style='thin', color="CBD5E1"), right=Side(style='thin', color="CBD5E1"), top=Side(style='thin', color="CBD5E1"), bottom=Side(style='thin', color="CBD5E1"))

        master_df = pd.concat([h["raw"] for h in hourly_data]).drop_duplicates(subset=[primary_col])
        
        def get_priority(pk):
            for h in hourly_data:
                if pk in h["sync"][primary_col].values: return 0
            return 1
            
        master_df['_priority'] = master_df[primary_col].apply(get_priority)
        master_df = master_df.sort_values(by=['_priority', primary_col])

        # Write Headers
        all_headers = [primary_col] + extra_cols + [h["display"] for h in hourly_data]
        for c, text in enumerate(all_headers, 1):
            cell = ws.cell(row=1, column=c, value=text)
            cell.font, cell.fill, cell.border = Font(bold=True, color="FFFFFF", size=11), header_fill, border
            cell.alignment = Alignment(horizontal="center")

        new_gp_counts = {i: 0 for i in range(len(hourly_data))}

        # Write Rows and Check for "NEW GPs"
        r_idx = 2
        for _, row_data in master_df.iterrows():
            pk = row_data[primary_col]
            ws.cell(row=r_idx, column=1, value=pk).border = border
            for i, col in enumerate(extra_cols, 2):
                ws.cell(row=r_idx, column=i, value=str(row_data.get(col, ""))).border = border

            for i, h_data in enumerate(hourly_data, len(extra_cols) + 2):
                cell = ws.cell(row=r_idx, column=i)
                cell.border = border
                list_idx = i - (len(extra_cols) + 2) 

                s_match = h_data["sync"][h_data["sync"][primary_col] == pk]
                u_match = h_data["unsync"][h_data["unsync"][primary_col] == pk]

                # --- NEW GP LOGIC ---
                is_newly_added = False
                if list_idx > 0: 
                    prev_h_data = hourly_data[list_idx - 1]
                    in_current = (not s_match.empty) or (not u_match.empty)
                    in_prev = pk in prev_h_data["raw"][primary_col].values
                    if in_current and not in_prev:
                        is_newly_added = True
                        new_gp_counts[list_idx] += 1 

                if not s_match.empty:
                    val = s_match.iloc[0][dt_col]
                    cell.value = val.strftime("%Y-%m-%d %H:%M") if pd.notnull(val) else "N/A"
                    cell.fill = new_gp_fill if is_newly_added else sync_fill
                elif not u_match.empty:
                    val = u_match.iloc[0][dt_col]
                    cell.value = val.strftime("%Y-%m-%d %H:%M") if pd.notnull(val) else "N/A"
                    cell.fill = new_gp_fill if is_newly_added else unsync_fill
                else:
                    cell.value = "-"
            r_idx += 1

        # Write Footer Totals
        r_idx += 1
        ws.cell(row=r_idx, column=1, value="SUCCESS COUNT (TODAY)").font = Font(bold=True)
        ws.cell(row=r_idx+1, column=1, value="STALE COUNT (OLD)").font = Font(bold=True)
        ws.cell(row=r_idx+2, column=1, value="NEW GP COUNT (ADDED)").font = Font(bold=True) 

        for i, h_data in enumerate(hourly_data, len(extra_cols) + 2):
            list_idx = i - (len(extra_cols) + 2)
            s_c = ws.cell(row=r_idx, column=i, value=len(h_data["sync"]))
            u_c = ws.cell(row=r_idx+1, column=i, value=len(h_data["unsync"]))
            n_c = ws.cell(row=r_idx+2, column=i, value=new_gp_counts[list_idx]) 
            
            for c in [s_c, u_c, n_c]:
                c.fill, c.font, c.border, c.alignment = total_fill, Font(bold=True), border, Alignment(horizontal="center")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output, 
            download_name="SYNC_ANALYTICS.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)